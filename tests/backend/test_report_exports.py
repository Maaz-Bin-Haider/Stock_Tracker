"""Export pipeline tests (FR-098…FR-101): jobs run through Celery (eager in
tests), files contain exactly the filtered dataset, and admin-only valuation
exports stay admin-only end to end."""

import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from apps.accounts.models import User
from apps.reports.models import ExportJob

pytestmark = pytest.mark.django_db

REPORTS = "/api/v1/reports/"
EXPORTS = "/api/v1/reports/exports/"


def create_export(client, key, format="XLSX", filters=None):
    return client.post(
        f"{REPORTS}{key}/export/",
        {"format": format, "filters": filters or {}},
        format="json",
    )


def download_bytes(client, job_id):
    response = client.get(f"{EXPORTS}{job_id}/download/")
    assert response.status_code == 200, getattr(response, "data", response)
    return b"".join(response.streaming_content)


class TestExcelExport:
    def test_xlsx_contains_filtered_dataset_only(self, report_world):
        response = create_export(
            report_world.admin_client,
            "purchase-report",
            filters={"location": report_world.sydney.pk},
        )
        assert response.status_code == 201, response.data
        assert response.data["status"] == "DONE"
        assert response.data["download_url"]

        content = download_bytes(report_world.admin_client, response.data["id"])
        sheet = load_workbook(io.BytesIO(content)).active
        cells = [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value]
        joined = " ".join(cells)
        assert "INV-P1" in joined
        assert "INV-P2" not in joined  # Dubai purchase filtered out (FR-098)
        assert "Sydney" in joined  # filter description printed on the sheet

    def test_valuation_summary_xlsx_has_all_sections(self, report_world):
        response = create_export(report_world.admin_client, "stock-valuation-summary")
        assert response.status_code == 201, response.data
        content = download_bytes(report_world.admin_client, response.data["id"])
        workbook = load_workbook(io.BytesIO(content))
        assert set(workbook.sheetnames) == {
            "Worth by Bucket",
            "Worth by Location",
            "Worth by Category",
            "Top Products by Value",
        }


class TestPdfExport:
    def test_pdf_bytes(self, report_world):
        response = create_export(report_world.admin_client, "gst-report", format="PDF")
        assert response.status_code == 201, response.data
        assert response.data["status"] == "DONE"
        content = download_bytes(report_world.admin_client, response.data["id"])
        assert content.startswith(b"%PDF")
        assert len(content) > 1000


class TestExportPermissions:
    def test_viewer_can_export_regular_reports(self, report_world, auth_client):
        viewer = auth_client(User.Role.VIEWER)
        response = create_export(viewer, "sales-report")
        assert response.status_code == 201, response.data
        assert download_bytes(viewer, response.data["id"]).startswith(b"PK")

    def test_valuation_export_admin_only(self, report_world, auth_client):
        for role in (User.Role.PURCHASE, User.Role.SALE, User.Role.VIEWER):
            response = create_export(auth_client(role), "stock-valuation-detail")
            assert response.status_code == 403, (role, response.data)
        assert not ExportJob.objects.filter(report_key="stock-valuation-detail").exists()

    def test_users_cannot_see_others_jobs(self, report_world, auth_client):
        admin_job = create_export(report_world.admin_client, "purchase-report").data
        viewer = auth_client(User.Role.VIEWER)
        assert viewer.get(f"{EXPORTS}{admin_job['id']}/").status_code == 404
        assert viewer.get(f"{EXPORTS}{admin_job['id']}/download/").status_code == 404
        # Admin sees everyone's jobs.
        listed = report_world.admin_client.get(EXPORTS)
        assert any(job["id"] == admin_job["id"] for job in listed.data["results"])

    def test_invalid_format_rejected(self, report_world):
        response = create_export(report_world.admin_client, "purchase-report", format="CSV")
        assert response.status_code == 400

    def test_invalid_filters_rejected_before_enqueue(self, report_world):
        response = create_export(
            report_world.admin_client, "purchase-report", filters={"date_from": "garbage"}
        )
        assert response.status_code == 400
        assert not ExportJob.objects.exists()

    def test_export_value_column_follows_role(self, report_world, auth_client):
        """total-company-stock exports carry the AED value only for admins."""
        viewer = auth_client(User.Role.VIEWER)
        viewer_job = create_export(viewer, "total-company-stock").data
        viewer_sheet = load_workbook(
            io.BytesIO(download_bytes(viewer, viewer_job["id"]))
        ).active
        viewer_cells = " ".join(
            str(cell.value) for row in viewer_sheet.iter_rows() for cell in row if cell.value
        )
        assert "Stock value (AED)" not in viewer_cells

        admin_job = create_export(report_world.admin_client, "total-company-stock").data
        admin_sheet = load_workbook(
            io.BytesIO(download_bytes(report_world.admin_client, admin_job["id"]))
        ).active
        admin_cells = " ".join(
            str(cell.value) for row in admin_sheet.iter_rows() for cell in row if cell.value
        )
        assert "Stock value (AED)" in admin_cells


class TestJobLifecycle:
    def test_failed_job_records_error(self, report_world, monkeypatch):
        from apps.reports import rendering

        def boom(*args, **kwargs):
            raise RuntimeError("renderer exploded")

        monkeypatch.setattr(rendering, "render_xlsx", boom)
        response = create_export(report_world.admin_client, "purchase-report")
        assert response.status_code == 201
        assert response.data["status"] == "FAILED"
        assert "renderer exploded" in response.data["error"]
        assert response.data["download_url"] is None

    def test_download_not_ready_rejected(self, report_world):
        job = ExportJob.objects.create(
            report_key="purchase-report",
            format=ExportJob.Format.XLSX,
            created_by=report_world.admin_client.user,
        )
        response = report_world.admin_client.get(f"{EXPORTS}{job.pk}/download/")
        assert response.status_code == 400

    def test_sale_totals_decimal(self, report_world):
        """Guard: report totals survive JSON rendering as numbers."""
        response = report_world.admin_client.get(f"{REPORTS}sales-report/")
        assert Decimal(str(response.data["totals"]["Quantity"])) == Decimal("3.00")
