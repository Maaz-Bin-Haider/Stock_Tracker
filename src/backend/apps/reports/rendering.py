"""Excel and PDF renderers for report exports (FR-099…FR-101).

Both renderers consume the same ``ReportResult`` the JSON endpoint serves, so
an export always matches what the user saw on screen with the same filters
(FR-098). PDF uses ReportLab with a clean, light, professional table theme
(FR-101). All datetimes are printed in Dubai business time (FR-128).
"""

import datetime
import io
import re
from decimal import Decimal

from apps.core.time import business_tz

from .definitions import Column, ReportResult

_HEADER_BG = "1E293B"  # slate-800
_HEADER_FG = "FFFFFF"
_STRIPE_BG = "F1F5F9"  # slate-100

NUMERIC_KINDS = {"qty", "money", "percent"}


def _display_value(value, kind: str):
    """Normalize a cell for rendering: Dubai-time datetimes, Yes/blank bools."""
    if value is None or value == "":
        return "" if kind not in NUMERIC_KINDS else value
    if kind == "bool":
        return "Yes" if value else ""
    if kind == "datetime" and isinstance(value, datetime.datetime):
        return value.astimezone(business_tz()).replace(tzinfo=None)
    return value


def _text_value(value, kind: str) -> str:
    value = _display_value(value, kind)
    if value is None or value == "":
        return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, datetime.date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, Decimal) and kind in ("money",):
        return f"{value:,.2f}"
    return str(value)


def _totals_line(totals: dict) -> str:
    return "   ".join(
        f"{label}: {_text_value(value, 'money' if isinstance(value, Decimal) else 'text')}"
        for label, value in totals.items()
    )


# ------------------------------------------------------------------- Excel


def render_xlsx(title: str, generated: str, filters_desc: str, result: ReportResult) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_font = Font(bold=True, color=_HEADER_FG)
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    used_titles: set[str] = set()

    for section in result.sections:
        # Excel forbids \ / ? * : [ ] in sheet names and caps them at 31 chars.
        safe_title = re.sub(r"[\\/?*:\[\]]", "-", section.title)
        sheet_title = safe_title[:31] or "Report"
        suffix = 2
        while sheet_title in used_titles:
            sheet_title = f"{safe_title[:28]} {suffix}"
            suffix += 1
        used_titles.add(sheet_title)
        sheet = workbook.create_sheet(sheet_title)

        sheet.append([title if len(result.sections) == 1 else f"{title} — {section.title}"])
        sheet.cell(row=1, column=1).font = Font(bold=True, size=13)
        sheet.append([generated])
        sheet.append([f"Filters — {filters_desc}"])
        if result.totals:
            sheet.append([_totals_line(result.totals)])
        sheet.append([])

        header_row = sheet.max_row + 1
        sheet.append([column.label for column in section.columns])
        for index, column in enumerate(section.columns, start=1):
            cell = sheet.cell(row=header_row, column=index)
            cell.font = header_font
            cell.fill = header_fill
            if column.kind in NUMERIC_KINDS:
                cell.alignment = Alignment(horizontal="right")

        widths = [len(column.label) for column in section.columns]
        for row in section.rows:
            values = []
            for index, column in enumerate(section.columns):
                value = _display_value(row.get(column.key), column.kind)
                values.append(value)
                widths[index] = min(max(widths[index], len(_text_value(value, column.kind))), 42)
            sheet.append(values)
            for index, column in enumerate(section.columns, start=1):
                cell = sheet.cell(row=sheet.max_row, column=index)
                if column.kind == "money":
                    cell.number_format = "#,##0.00"
                elif column.kind == "qty":
                    cell.number_format = "#,##0.##"
                elif column.kind == "date":
                    cell.number_format = "DD/MM/YYYY"
                elif column.kind == "datetime":
                    cell.number_format = "DD/MM/YYYY HH:MM"

        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width + 3
        sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------- PDF


def render_pdf(title: str, generated: str, filters_desc: str, result: ReportResult) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    header_bg = colors.HexColor(f"#{_HEADER_BG}")
    stripe_bg = colors.HexColor(f"#{_STRIPE_BG}")
    grid = colors.HexColor("#CBD5E1")
    body_color = colors.HexColor("#0F172A")

    title_style = ParagraphStyle(
        "ReportTitle", fontName="Helvetica-Bold", fontSize=15, textColor=body_color
    )
    meta_style = ParagraphStyle(
        "Meta", fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#475569")
    )
    section_style = ParagraphStyle(
        "SectionTitle",
        fontName="Helvetica-Bold",
        fontSize=11,
        textColor=body_color,
        spaceBefore=8,
    )
    cell_style = ParagraphStyle(
        "Cell", fontName="Helvetica", fontSize=7, leading=9, textColor=body_color
    )
    header_style = ParagraphStyle(
        "HeaderCell", fontName="Helvetica-Bold", fontSize=7, leading=9, textColor=colors.white
    )

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
    )
    available_width = document.width

    story = [
        Paragraph(title, title_style),
        Spacer(1, 4),
        Paragraph(generated, meta_style),
        Paragraph(f"Filters — {filters_desc}", meta_style),
    ]
    if result.totals:
        story.append(Paragraph(_totals_line(result.totals), meta_style))
    story.append(Spacer(1, 6))

    def _column_widths(section_columns: list[Column], rows: list[dict]) -> list[float]:
        # Content-aware widths: numeric cells never wrap, so their columns
        # must fit the longest value; text columns get the remaining space.
        weights = []
        for column in section_columns:
            longest = max(
                (len(_text_value(row.get(column.key), column.kind)) for row in rows),
                default=4,
            )
            if column.kind in NUMERIC_KINDS:
                weights.append(min(max(longest + 1, 5), 14))
            else:
                header_word = max((len(word) for word in column.label.split()), default=6)
                weights.append(min(max(longest, header_word, 6), 26))
        total = sum(weights)
        return [available_width * weight / total for weight in weights]

    for section in result.sections:
        if len(result.sections) > 1:
            story.append(Paragraph(section.title, section_style))
            story.append(Spacer(1, 3))
        data = [
            [Paragraph(column.label, header_style) for column in section.columns]
        ]
        for row in section.rows:
            cells = []
            for column in section.columns:
                text = _text_value(row.get(column.key), column.kind)
                # Plain strings for numeric cells so the table's RIGHT align
                # applies; Paragraphs (which wrap) for free text.
                cells.append(text if column.kind in NUMERIC_KINDS else Paragraph(text, cell_style))
            data.append(cells)
        if not section.rows:
            data.append(
                [Paragraph("No data for the selected filters.", cell_style)]
                + [Paragraph("", cell_style)] * (len(section.columns) - 1)
            )

        table = Table(
            data, colWidths=_column_widths(section.columns, section.rows), repeatRows=1
        )
        style = [
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("TEXTCOLOR", (0, 1), (-1, -1), body_color),
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("GRID", (0, 0), (-1, -1), 0.4, grid),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, stripe_bg]),
        ]
        for index, column in enumerate(section.columns):
            if column.kind in NUMERIC_KINDS:
                style.append(("ALIGN", (index, 0), (index, -1), "RIGHT"))
        table.setStyle(TableStyle(style))
        story.append(table)
        story.append(Spacer(1, 10))

    document.build(story)
    return buffer.getvalue()
